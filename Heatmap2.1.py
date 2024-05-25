#import
import tkinter as tk
from tkinter import Tk,filedialog,Button
from openpyxl import load_workbook
import auto_py_to_exe
#AbcTo123
def AbcTo123 (letters):
    abcpickup=str(letters).upper()
    if abcpickup=="A" :
        abcpickuped="1"
    elif abcpickup=="B" :
        abcpickuped="2"
    elif abcpickup=="C" :
        abcpickuped="3"
    elif abcpickup=="D" :
        abcpickuped="4"
    elif abcpickup=="E" :
        abcpickuped="5"
    elif abcpickup=="F" :
        abcpickuped="6"
    elif abcpickup=="G" :
        abcpickuped="7"
    elif abcpickup=="H" :
        abcpickuped="8"
    elif abcpickup=="I" :
        abcpickuped="9"
    elif abcpickup=="J" :
        abcpickuped="10"
    elif abcpickup=="K" :
        abcpickuped="11"
    elif abcpickup=="L" :
        abcpickuped="12"
    elif abcpickup=="M" :
        abcpickuped="13"
    elif abcpickup=="N" :
        abcpickuped="14"
    elif abcpickup=="O" :
        abcpickuped="15"
    elif abcpickup=="P" :
        abcpickuped="16"
    elif abcpickup=="Q" :
        abcpickuped="17"
    elif abcpickup=="R" :
        abcpickuped="18"
    elif abcpickup=="S" :
        abcpickuped="19"
    elif abcpickup=="T" :
        abcpickuped="20"
    elif abcpickup=="U" :
        abcpickuped="21"
    elif abcpickup=="V" :
        abcpickuped="22"
    elif abcpickup=="W" :
        abcpickuped="23"
    elif abcpickup=="X" :
        abcpickuped="24"
    elif abcpickup=="Y" :
        abcpickuped="25"
    elif abcpickup=="Z" :
        abcpickuped="26"
    return abcpickuped
#LetterToNumber
def LetterToNumber(letter):
    for i in range (1,len(letter)+1):
        i2=int(len(letter))+1-i
        if i==int(len(letter)):
            globals()["the"+str(i)+"number"]=int(AbcTo123(letter[(i-1)]))
        else:
            globals()["the"+str(i)+"number"]=1
            for i3 in range (1,i2):
                globals()["the"+str(i)+"number"]=globals()["the"+str(i)+"number"]*26
            globals()["the"+str(i)+"number"]=int(globals()["the"+str(i)+"number"]*int(AbcTo123(letter[(i-1)])))
    number=int(0)
    for i in range (1,len(letter)+1):
        number=number+int(globals()["the"+str(i)+"number"])
    return number
#RgbToHex
def RgbToHex(Rgb):
    if Rgb==0:
        Hex="F"
    elif Rgb==1:
        Hex="E"
    elif Rgb==2:
        Hex="D"
    elif Rgb==3:
        Hex="C"
    elif Rgb==4:
        Hex="B"
    elif Rgb==5:
        Hex="A"
    elif Rgb==6:
        Hex="9"
    elif Rgb==7:
        Hex="8"
    elif Rgb==8:
        Hex="7"
    elif Rgb==9:
        Hex="6"
    elif Rgb==10:
        Hex="5"
    elif Rgb==11:
        Hex="4"
    elif Rgb==12:
        Hex="3"
    elif Rgb==13:
        Hex="2"
    elif Rgb==14:
        Hex="1"
    elif Rgb==15:
        Hex="0"
    return Hex
#Count
def Count():
    #GetTrueValue
    AskLocation1ColAnsStr=str(AskLocation1Col.get())
    AskLocation1RowAnsInt=int(AskLocation1Row.get())
    AskLocation2ColAnsStr=str(AskLocation2Col.get())
    AskLocation2RowAnsInt=int(AskLocation2Row.get())
    #AbcTo123
    AskLocation1ColAnsInt=int(LetterToNumber(AskLocation1ColAnsStr))
    AskLocation2ColAnsInt=int(LetterToNumber(AskLocation2ColAnsStr))
    #LoadWorkBook
    Wb=load_workbook(globals()["Filelocation"])
    Ws=Wb.active
    #CountValueIntoColor
    Biggest=0
    Smallest=0
    for Row in range (AskLocation1RowAnsInt,AskLocation2RowAnsInt+1):
        for Col in range (AskLocation1ColAnsInt,AskLocation2ColAnsInt+1):
            globals()["Col"+str(Col)+"Row"+str(Row)+"Value"]=float((Ws.cell(row=Row,column=Col)).value)
            if globals()["Col"+str(Col)+"Row"+str(Row)+"Value"]>Biggest:
                Biggest=globals()["Col"+str(Col)+"Row"+str(Row)+"Value"]
            if globals()["Col"+str(Col)+"Row"+str(Row)+"Value"]<Smallest:
                Smallest=globals()["Col"+str(Col)+"Row"+str(Row)+"Value"]
    R=float(float(225)/(Biggest-Smallest))
    for Row in range (AskLocation1RowAnsInt,AskLocation2RowAnsInt+1):
        for Col in range (AskLocation1ColAnsInt,AskLocation2ColAnsInt+1):
            globals()["Col"+str(Col)+"Row"+str(Row)+"Rgb"]=R*(globals()["Col"+str(Col)+"Row"+str(Row)+"Value"]-Smallest)
            Hex10=RgbToHex(int(globals()["Col"+str(Col)+"Row"+str(Row)+"Rgb"]//16))
            Hex1=RgbToHex(int(globals()["Col"+str(Col)+"Row"+str(Row)+"Rgb"]%16))
            globals()["Col"+str(Col)+"Row"+str(Row)+"Hex"]=str("#FF"+str(Hex10)+str(Hex1)+str(Hex10)+str(Hex1))
    #EndWindow
    Height=str((AskLocation2RowAnsInt-AskLocation1RowAnsInt+1)*10+20)
    Width=str((AskLocation2ColAnsInt-AskLocation1ColAnsInt+1)*10+20)
    Size=str(Width+"x"+Height)
    end=tk.Tk()
    end.title("結果")
    canvas=tk.Canvas(end,width=Width,height=Height,bg='white')
    canvas.place(x=0,y=0)
    #ShowEnd
    for Row in range (AskLocation1RowAnsInt,AskLocation2RowAnsInt+1):
        for Col in range (AskLocation1ColAnsInt,AskLocation2ColAnsInt+1):
            canvas.create_rectangle((Col-AskLocation1ColAnsInt)*10+10,(Row-AskLocation1RowAnsInt)*10+10,(Col-AskLocation1ColAnsInt)*10+20,(Row-AskLocation1RowAnsInt)*10+20,fill=globals()["Col"+str(Col)+"Row"+str(Row)+"Hex"],width=0)
    #ShowEndWindow
    end.geometry(Size)
    end.mainloop()
def Browesfile():
    globals()["Filelocation"]=filedialog.askopenfilename(initialdir="your directory path", title="選取.xlsx檔案",filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
#Asklocation
##Window
SetUp=tk.Tk()
SetUp.geometry("430x190")
SetUp.title("設定")
##labal
AskFileLocation1Test=tk.Label(SetUp,text="選取檔案:",font=("Times",20))
AskFileLocation1Test.place(x=10,y=10)
AskLocation1Test=tk.Label(SetUp,text="最左上角的座標=>欄:        列:",font=("Times",20))
AskLocation1Test.place(x=10,y=50)
AskLocation2Text=tk.Label(SetUp,text="最右下角的座標=>欄:        列:",font=("Times",20))
AskLocation2Text.place(x=10,y=90)
##Entry
AskLocation1Col=tk.Entry(SetUp,font=("Times",20),width=3)
AskLocation1Col.place(x=280,y=50)
AskLocation1Row=tk.Entry(SetUp,font=("Times",20),width=3)
AskLocation1Row.place(x=370,y=50)
AskLocation2Col=tk.Entry(SetUp,font=("Times",20),width=3)
AskLocation2Col.place(x=280,y=90)
AskLocation2Row=tk.Entry(SetUp,font=("Times",20),width=3)
AskLocation2Row.place(x=370,y=90)
##Enter
AskLocationEnter=tk.Button(SetUp,text="瀏覽",font=("Times",12),background=("#1fcaff"),command=Browesfile)
AskLocationEnter.place(x=140,y=10)
AskLocationEnter=tk.Button(SetUp,text="確認",font=("Times",20),background=("#1fcaff"),command=Count)
AskLocationEnter.place(x=10,y=130)
##Show
SetUp.mainloop()